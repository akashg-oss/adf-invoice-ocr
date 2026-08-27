import io
import re
import fitz
import pandas as pd
import streamlit as st
from rapidfuzz import fuzz, process

st.set_page_config(
    page_title="ADF Invoice OCR → Standard Excel",
    page_icon="📄",
    layout="wide",
)

OUTPUT_COLUMNS = [
    "SKU Code",
    "EAN",
    "Name of FG",
    "Invoice Number",
    "City",
    "PO Number",
    "Quantity Dispatched (PCS)",
    "Price of FG (₹/PCS)",
]


# ============================================================
# HELPERS
# ============================================================

def clean_text(value):
    if value is None or pd.isna(value):
        return ""
    return str(value).strip()


def norm(value):
    """Normalize text for matching."""
    value = clean_text(value).upper()
    value = value.replace("’", "'")
    value = re.sub(r"[^A-Z0-9]+", " ", value)
    return re.sub(r"\s+", " ", value).strip()


def number(value):
    """Convert Indian-formatted numbers such as 2,61,465.60 to float."""
    if value is None:
        return 0.0
    return float(str(value).replace(",", "").strip())


# ============================================================
# MASTER SKU FILE
# ============================================================

def load_master(file):
    """
    Reads the user's master Excel.

    It automatically detects the header row, so both of these work:

        Row 1: EAN | SKU Code | SKU Name

    or

        Row 1: blank
        Row 2: EAN | SKU Code | SKU Name
    """
    raw = pd.read_excel(file, header=None, dtype=str)

    header_row = None

    for i in range(min(len(raw), 30)):
        row = [norm(x) for x in raw.iloc[i].tolist()]

        has_ean = any(
            x in row
            for x in ["EAN", "EAN CODE", "EAN NO", "EAN NUMBER"]
        )
        has_sku = any(
            x in row
            for x in ["SKU", "SKU CODE", "PRODUCT SKU"]
        )
        has_name = any(
            x in row
            for x in [
                "SKU NAME",
                "NAME",
                "PRODUCT NAME",
                "NAME OF FG",
                "FG NAME",
                "DESCRIPTION",
            ]
        )

        if has_ean and has_sku and has_name:
            header_row = i
            break

    if header_row is None:
        raise ValueError(
            "Could not find the master headers. "
            "The Excel must contain EAN, SKU Code and SKU Name."
        )

    headers = [
        clean_text(x) for x in raw.iloc[header_row].tolist()
    ]

    data = raw.iloc[header_row + 1:].copy()
    data.columns = headers

    # Remove blank/duplicate-looking columns.
    data = data.loc[:, [c != "" for c in data.columns]]

    normalized_headers = {norm(c): c for c in data.columns}

    def find_column(options):
        for option in options:
            if norm(option) in normalized_headers:
                return normalized_headers[norm(option)]

        for col in data.columns:
            col_norm = norm(col)
            for option in options:
                option_norm = norm(option)
                if option_norm in col_norm or col_norm in option_norm:
                    return col

        return None

    ean_col = find_column(
        ["EAN", "EAN Code", "EAN No", "EAN Number"]
    )
    sku_col = find_column(
        ["SKU Code", "SKU", "Product SKU"]
    )
    name_col = find_column(
        [
            "SKU Name",
            "Name",
            "Product Name",
            "Name of FG",
            "FG Name",
            "Description",
        ]
    )

    if not all([ean_col, sku_col, name_col]):
        raise ValueError(
            "Master Excel must contain these columns: "
            "EAN, SKU Code and SKU Name."
        )

    master = data[[ean_col, sku_col, name_col]].copy()
    master.columns = ["EAN", "SKU Code", "SKU Name"]

    for col in master.columns:
        master[col] = (
            master[col]
            .fillna("")
            .astype(str)
            .str.replace(r"\.0$", "", regex=True)
            .str.strip()
        )

    master = master[
        (master["EAN"] != "")
        & (master["SKU Code"] != "")
        & (master["SKU Name"] != "")
    ].drop_duplicates()

    return master


# ============================================================
# PDF TEXT EXTRACTION
# ============================================================

def get_pdf_text(pdf_bytes):
    """
    Use PyMuPDF first because it preserves the ADF invoice text/table
    better than the earlier parser.

    OCR is used only if the PDF has no usable text layer.
    """
    text = ""

    try:
        doc = fitz.open(stream=pdf_bytes, filetype="pdf")
        text = "\n".join(
            page.get_text("text") or ""
            for page in doc
        )
        doc.close()
    except Exception:
        text = ""

    if len(re.sub(r"\s+", "", text)) >= 250:
        return text

    # OCR fallback
    try:
        import pytesseract
        from PIL import Image

        doc = fitz.open(stream=pdf_bytes, filetype="pdf")
        pages = []

        for page in doc:
            pix = page.get_pixmap(
                matrix=fitz.Matrix(2.5, 2.5),
                alpha=False
            )
            image = Image.open(
                io.BytesIO(pix.tobytes("png"))
            )
            pages.append(pytesseract.image_to_string(image))

        doc.close()
        return "\n".join(pages)

    except Exception:
        return text


# ============================================================
# INVOICE HEADER
# ============================================================

def extract_invoice_number(text):
    patterns = [
        r"\b(ADF/\d{4}-\d{2}/\d+)\b",
        r"Invoice\s*No\.?\s*[:\n\s]*([A-Z0-9/\-]+)",
    ]

    for pattern in patterns:
        match = re.search(pattern, text, re.I)
        if match:
            return match.group(1).strip()

    return ""


def extract_po_number(text):
    patterns = [
        r"Buyer[’']?s\s+Order\s+No\.?\s*(?:Dated\s+)?([0-9]{8,12})",
        r"Buyer[’']?s\s+Order\s+No\.?.{0,80}?([0-9]{8,12})",
        r"\b(4\d{9})\b",
        r"PO\s*(?:No|Number)?\.?\s*[:\-]?\s*([0-9]{8,12})",
    ]

    for pattern in patterns:
        match = re.search(pattern, text, re.I | re.S)
        if match:
            return match.group(1).strip()

    return ""


def extract_city(text):
    """
    Extract the city to which the invoice is addressed.

    Priority:
    1. Explicit Destination field.
    2. City in the Consignee (Ship to) address.

    Known ADF spelling correction:
    BANGLORE → Bangalore
    """
    destination_match = re.search(
        r"Destination\s*[:\n\s]*([A-Za-z]+)",
        text,
        re.I,
    )

    destination = (
        destination_match.group(1).strip()
        if destination_match
        else ""
    )

    d = norm(destination)

    if d in {"BANGLORE", "BANGALORE", "BENGALURU"}:
        return "Bangalore"

    if d == "THANE":
        return "Thane"

    if d == "KOLKATA":
        return "Kolkata"

    if d == "HOWRAH":
        return "Howrah"

    # Haryana is a state in these ADF invoices; the consignee address
    # contains Gurgaon, which is the city we want in the output.
    if d == "HARYANA":
        city_match = re.search(
            r"(?:Gurgaon|Gurugram)\s*[-–]?\s*\d{6}",
            text,
            re.I,
        )
        if city_match:
            return re.sub(
                r"\s*[-–]?\s*\d{6}",
                "",
                city_match.group(0),
                flags=re.I,
            ).strip().title()

    # Generic fallback: look for a city immediately before a 6-digit PIN.
    pin_match = re.search(
        r"([A-Za-z][A-Za-z .'-]{2,40})\s*[-–]\s*\d{6}\b",
        text,
        re.I,
    )
    if pin_match:
        candidate = pin_match.group(1).strip()
        candidate = re.sub(
            r"^(District|Village|Tal|Tehsil)\s+",
            "",
            candidate,
            flags=re.I,
        )
        return candidate.title()

    return destination


# ============================================================
# ADF PRODUCT LINE EXTRACTION
# ============================================================

def extract_invoice_items(text):
    """
    ADF invoice product rows look like:

    1 FG-PURPLLE-PER-20ML
      -STAY-OUD TILL DOWN X 48
      1,63,416.00 PCS 61.90 2,640.0000 PCS 55.00 BOX 33030050

    PyMuPDF may split the description over several lines.
    We therefore normalize whitespace first.

    We extract:
      - invoice FG description
      - pack size
      - shipped PCS
      - price per PCS
    """

    clean = re.sub(r"\s+", " ", text)

    # Capture the complete FG description through the X pack-size marker,
    # then capture amount/rate/quantity from the table row.
    pattern = re.compile(
        r"""
        (FG-PURPLLE-PER-(?:20|50|100)ML\s*-\s*.+?)
        \s+X\s+(\d+)
        \s+
        ([\d,]+\.\d+)
        \s+PCS
        \s+
        ([\d,]+\.\d+)
        \s+
        ([\d,]+\.\d+)
        \s+PCS
        \s+
        [\d,]+\.\d+
        \s+BOX
        \s+
        33030050
        """,
        re.I | re.X,
    )

    items = []

    for match in pattern.finditer(clean):
        description = re.sub(
            r"\s+",
            " ",
            match.group(1).strip(),
        )

        pack_size = int(match.group(2))
        billed_amount = number(match.group(3))
        rate = number(match.group(4))
        quantity_pcs = int(
            round(number(match.group(5)))
        )

        # If the explicit rate exists, use it.
        # Otherwise calculate it from the invoice amount.
        if rate <= 0 and quantity_pcs > 0:
            rate = round(
                billed_amount / quantity_pcs,
                2,
            )

        items.append(
            {
                "description": description,
                "pack_size": pack_size,
                "quantity": quantity_pcs,
                "rate": rate,
                "amount": billed_amount,
            }
        )

    # More tolerant fallback if a future ADF PDF changes column spacing.
    if not items:
        fallback = re.compile(
            r"""
            (FG-PURPLLE-PER-\d+ML\s*-\s*.+?)
            \s+X\s+(\d+)
            .*?
            ([\d,]+\.\d+)
            \s+PCS
            \s+
            ([\d,]+\.\d+)
            \s+
            ([\d,]+\.\d+)
            \s+PCS
            """,
            re.I | re.X,
        )

        for match in fallback.finditer(clean):
            description = re.sub(
                r"\s+",
                " ",
                match.group(1).strip(),
            )

            quantity = int(
                round(number(match.group(5)))
            )
            rate = number(match.group(4))
            amount = number(match.group(3))

            items.append(
                {
                    "description": description,
                    "pack_size": int(match.group(2)),
                    "quantity": quantity,
                    "rate": rate,
                    "amount": amount,
                }
            )

    # Remove duplicates.
    unique = []
    seen = set()

    for item in items:
        key = (
            norm(item["description"]),
            item["quantity"],
            round(item["rate"], 2),
        )

        if key not in seen:
            seen.add(key)
            unique.append(item)

    return unique


# ============================================================
# MASTER SKU MATCHING
# ============================================================

def product_key(value):
    """
    Build a typo-tolerant product key.

    ADF descriptions can contain:
      - spelling errors (ECSTECY instead of ECSTASY)
      - split words (LOVE STRUCK vs LOVESTRUCK)
      - split words (DAY DREAMS vs DAYDREAMS)
      - small OCR errors

    We therefore:
      1. Keep the volume separately.
      2. Remove brand/generic perfume wording.
      3. Remove spaces/hyphens for comparison.
      4. Compare the resulting product name using fuzzy matching.
    """
    s = norm(value)

    volume_match = re.search(
        r"\b(20ML|50ML|100ML)\b",
        s,
    )
    volume = volume_match.group(1) if volume_match else ""

    # Remove generic words that do not identify the product.
    remove_words = [
        "FACES",
        "CANADA",
        "EAU",
        "DE",
        "PARFUM",
        "MINI",
        "PURPLLE",
        "PER",
        "FG",
    ]

    for word in remove_words:
        s = re.sub(
            rf"\b{re.escape(word)}\b",
            " ",
            s,
        )

    # Remove volume from the product-name portion; it is handled separately.
    s = re.sub(
        r"\b(?:20ML|50ML|100ML)\b",
        " ",
        s,
    )

    # Make common OCR/spelling variants consistent.
    replacements = {
        "ECSTECY": "ECSTASY",
        "ROMENTIC": "ROMANTIC",
        "LOVE STRUCK": "LOVESTRUCK",
        "DAY DREAMS": "DAYDREAMS",
        "LOVE-STRUCK": "LOVESTRUCK",
        "DAY-DREAMS": "DAYDREAMS",
    }

    for old, new in replacements.items():
        s = s.replace(old, new)

    # Spaces/hyphens should not matter:
    # LOVE STRUCK DELIGHT == LOVESTRUCK DELIGHT
    compact = re.sub(r"[^A-Z0-9]", "", s)

    return compact, volume


def match_to_master(invoice_description, master):
    """
    Match invoice product to the master SKU.

    Important:
    - Volume must match (20ml cannot map to 100ml).
    - Fuzzy confidence threshold is 85%.
    - The best 85%+ match is accepted even when the invoice has
      spelling/OCR mistakes.
    """

    invoice_key, invoice_volume = product_key(invoice_description)

    candidates = []

    for index, row in master.iterrows():
        master_key, master_volume = product_key(
            row["SKU Name"]
        )

        # Never map a 20ml item to a 50/100ml SKU.
        if (
            invoice_volume
            and master_volume
            and invoice_volume != master_volume
        ):
            continue

        candidates.append(
            (index, master_key)
        )

    if not candidates:
        return None, 0

    candidate_keys = [
        key for _, key in candidates
    ]

    # Exact normalized match.
    for (index, key) in candidates:
        if key == invoice_key:
            return master.loc[index], 100

    # Fuzzy matching after removing spacing/hyphen differences.
    best = process.extractOne(
        invoice_key,
        candidate_keys,
        scorer=fuzz.ratio,
    )

    if best:
        _, score, candidate_position = best
        master_index = candidates[candidate_position][0]

        if score >= 85:
            return (
                master.loc[master_index],
                score,
            )

    return None, 0


# ============================================================
# PROCESS ONE INVOICE
# ============================================================

def process_invoice(file, master):
    pdf_bytes = file.read()
    text = get_pdf_text(pdf_bytes)

    invoice_number = extract_invoice_number(text)
    po_number = extract_po_number(text)
    city = extract_city(text)
    items = extract_invoice_items(text)

    rows = []
    issues = []

    if not invoice_number:
        issues.append(
            f"{file.name}: Invoice number not detected"
        )

    if not po_number:
        issues.append(
            f"{file.name}: PO number not detected"
        )

    if not city:
        issues.append(
            f"{file.name}: City not detected"
        )

    if not items:
        issues.append(
            f"{file.name}: No invoice line items detected"
        )

    for item in items:
        matched, score = match_to_master(
            item["description"],
            master,
        )

        if matched is None:
            rows.append(
                [
                    "",
                    "",
                    item["description"],
                    invoice_number,
                    city,
                    po_number,
                    item["quantity"],
                    item["rate"],
                ]
            )

            issues.append(
                f"{file.name}: SKU mapping not found → "
                f"{item['description']}"
            )

        else:
            rows.append(
                [
                    matched["SKU Code"],
                    matched["EAN"],
                    matched["SKU Name"],
                    invoice_number,
                    city,
                    po_number,
                    item["quantity"],
                    item["rate"],
                ]
            )

            if score < 95:
                issues.append(
                    f"{file.name}: Fuzzy SKU match {score:.0f}% → "
                    f"{item['description']} = "
                    f"{matched['SKU Name']}"
                )

    return rows, issues


# ============================================================
# EXCEL CREATION
# ============================================================

def create_excel(result):
    output = io.BytesIO()

    result = result.copy()

    result["Quantity Dispatched (PCS)"] = pd.to_numeric(
        result["Quantity Dispatched (PCS)"],
        errors="coerce",
    ).astype("Int64")

    result["Price of FG (₹/PCS)"] = pd.to_numeric(
        result["Price of FG (₹/PCS)"],
        errors="coerce",
    )

    # Summary by invoice.
    calc = result.copy()
    calc["_FG Value"] = (
        calc["Quantity Dispatched (PCS)"]
        * calc["Price of FG (₹/PCS)"]
    )

    summary = (
        calc.groupby(
            ["Invoice Number", "City", "PO Number"],
            dropna=False,
            as_index=False,
        )
        .agg(
            **{
                "Total Quantity Dispatched (PCS)": (
                    "Quantity Dispatched (PCS)",
                    "sum",
                ),
                "Total FG Value (₹)": (
                    "_FG Value",
                    "sum",
                ),
            }
        )
    )

    with pd.ExcelWriter(
        output,
        engine="openpyxl",
    ) as writer:

        result[OUTPUT_COLUMNS].to_excel(
            writer,
            index=False,
            sheet_name="Invoice Details",
        )

        summary.to_excel(
            writer,
            index=False,
            sheet_name="Summary",
        )

        # Basic formatting.
        from openpyxl.styles import Font, PatternFill

        for sheet_name in [
            "Invoice Details",
            "Summary",
        ]:
            ws = writer.book[sheet_name]

            ws.freeze_panes = "A2"
            ws.auto_filter.ref = ws.dimensions

            for cell in ws[1]:
                cell.font = Font(
                    bold=True,
                    color="FFFFFF",
                )
                cell.fill = PatternFill(
                    "solid",
                    fgColor="1F4E78",
                )

            for column in ws.columns:
                max_len = max(
                    len(str(cell.value or ""))
                    for cell in column
                )
                ws.column_dimensions[
                    column[0].column_letter
                ].width = min(
                    max(max_len + 2, 12),
                    60,
                )

    return output.getvalue()


# ============================================================
# STREAMLIT UI
# ============================================================

st.title("ADF Invoice OCR → Standard Excel")

st.write(
    "Upload the EAN/SKU master and ADF invoice PDFs. "
    "The app will always generate the same standard Excel format."
)

st.info(
    "Master Excel format: **EAN | SKU Code | SKU Name**"
)

master_file = st.file_uploader(
    "1. Upload EAN ↔ SKU Master Excel",
    type=["xlsx", "xls"],
)

invoice_files = st.file_uploader(
    "2. Upload ADF Invoice PDFs",
    type=["pdf"],
    accept_multiple_files=True,
)

if master_file:
    try:
        master_preview = load_master(master_file)
        st.success(
            f"Master loaded successfully: "
            f"{len(master_preview)} SKU mappings."
        )
    except Exception as error:
        st.error(str(error))
        st.stop()

if master_file and invoice_files:

    if st.button(
        "🚀 Process Invoices",
        type="primary",
    ):
        master_file.seek(0)
        master = load_master(master_file)

        all_rows = []
        all_issues = []

        progress = st.progress(0)

        for i, invoice_file in enumerate(
            invoice_files,
            start=1,
        ):
            rows, issues = process_invoice(
                invoice_file,
                master,
            )

            all_rows.extend(rows)
            all_issues.extend(issues)

            progress.progress(
                i / len(invoice_files)
            )

        if not all_rows:
            st.error(
                "No line items extracted from the invoice PDFs."
            )

            st.info(
                "The master file is being read correctly. "
                "The remaining issue is the invoice PDF format."
            )

            if all_issues:
                st.warning("\n".join(all_issues))

        else:
            result = pd.DataFrame(
                all_rows,
                columns=OUTPUT_COLUMNS,
            )

            st.success(
                f"Processed {len(invoice_files)} invoice(s) "
                f"and extracted {len(result)} line item(s)."
            )

            st.dataframe(
                result,
                use_container_width=True,
            )

            excel_data = create_excel(result)

            st.download_button(
                label="📥 Download Standardized Excel",
                data=excel_data,
                file_name="ADF_Invoice_Dispatch_Details.xlsx",
                mime=(
                    "application/vnd.openxmlformats-officedocument."
                    "spreadsheetml.sheet"
                ),
            )

            if all_issues:
                st.warning(
                    f"{len(all_issues)} item(s) need review."
                )

                st.dataframe(
                    pd.DataFrame(
                        {"Review / Issue": all_issues}
                    ),
                    use_container_width=True,
                )
            else:
                st.success(
                    "✅ All invoice line items were mapped successfully."
                )
